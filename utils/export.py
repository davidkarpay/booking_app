"""
Export functions for the PBSO Booking Blotter
"""
import csv
import json
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt5.QtWidgets import QFileDialog
from logger import logger

def export_to_csv(booking_data, parent=None):
    """
    Export booking data to a CSV file
    
    Args:
        booking_data: List of booking record dictionaries
        parent: Parent widget for dialog (optional)
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    if not booking_data:
        logger.warning("No data to export")
        return False
            
    file_path, _ = QFileDialog.getSaveFileName(
        parent, "Save CSV File", "", "CSV Files (*.csv);;All Files (*)"
    )
    
    if not file_path:
        return False
        
    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            # Get all possible field names from all records
            fieldnames = set()
            for record in booking_data:
                fieldnames.update(record.keys())
            
            # Sort field names for consistent column order
            fieldnames = sorted(fieldnames)
            
            # Remove Raw Data field if present (too large for CSV)
            if "Raw Data" in fieldnames:
                fieldnames.remove("Raw Data")
            
            # Create CSV writer and write header
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            # Write data rows
            for record in booking_data:
                # Create a copy without Raw Data
                record_copy = {k: v for k, v in record.items() if k != "Raw Data"}
                writer.writerow(record_copy)
        
        logger.info(f"Exported {len(booking_data)} records to CSV: {file_path}")
        return True
    except Exception as e:
        logger.error(f"CSV export error: {str(e)}")
        return False

def export_to_excel(booking_data, parent=None):
    """
    Export booking data to an Excel file with formatting
    
    Args:
        booking_data: List of booking record dictionaries
        parent: Parent widget for dialog (optional)
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    if not booking_data:
        logger.warning("No data to export")
        return False
            
    file_path, _ = QFileDialog.getSaveFileName(
        parent, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)"
    )
    
    if not file_path:
        return False
        
    try:
        # Create a copy of data without Raw Data field (too large for Excel)
        clean_data = []
        for record in booking_data:
            record_copy = {k: v for k, v in record.items() if k != "Raw Data"}
            clean_data.append(record_copy)
        
        # Convert data to DataFrame
        df = pd.DataFrame(clean_data)
        
        # Create Excel writer
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        # Write data to Excel
        df.to_excel(writer, sheet_name='Booking Results', index=False)
        
        # Access the workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets['Booking Results']
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for i, col in enumerate(df.columns):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(i + 1)
            
            # Find the maximum length in the column
            for j in range(1, len(df) + 2):  # +2 for header and 1-based indexing
                cell_value = str(worksheet.cell(row=j, column=i + 1).value)
                max_length = max(max_length, len(cell_value))
            
            # Set the column width (with some padding)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add a summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        
        # Calculate summary statistics
        in_custody_count = sum(1 for item in booking_data if item.get("Status") == "In Custody")
        released_count = sum(1 for item in booking_data if item.get("Status") == "Released")
        
        # Get time served info
        days_served = [item.get("Time Served (Days)", 0) for item in booking_data 
                      if isinstance(item.get("Time Served (Days)", 0), (int, float))]
        
        if days_served:
            avg_days = sum(days_served) / len(days_served)
            max_days = max(days_served)
            min_days = min(days_served)
        else:
            avg_days = max_days = min_days = 0
        
        # Add summary data
        summary_data = [
            ["Summary Statistics", ""],
            ["Total Records Found", len(booking_data)],
            ["Currently In Custody", in_custody_count],
            ["Released", released_count],
            ["Average Time Served (days)", round(avg_days, 1)],
            ["Longest Time Served (days)", max_days],
            ["Shortest Time Served (days)", min_days],
            ["Report Generated", datetime.now().strftime("%m/%d/%Y %H:%M")],
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, cell_value in enumerate(row_data, 1):
                summary_sheet.cell(row=row_num, column=col_num).value = cell_value
        
        # Format the summary sheet
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 15
        
        # Save the Excel file
        writer.close()
        
        logger.info(f"Exported {len(booking_data)} records to Excel: {file_path}")
        return True
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def export_filtered_data(visible_records, parent=None):
    """
    Export only the filtered/visible data
    
    Args:
        visible_records: List of visible booking record dictionaries
        parent: Parent widget for dialog (optional)
        
    Returns:
        bool: True if export was successful, False otherwise
    """
    if not visible_records:
        logger.warning("No visible records to export")
        return False
            
    file_path, _ = QFileDialog.getSaveFileName(
        parent, "Save Filtered Data", "", "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)"
    )
    
    if not file_path:
        return False
        
    try:
        if file_path.lower().endswith('.csv'):
            return export_to_csv(visible_records, parent)
        elif file_path.lower().endswith('.xlsx'):
            return export_to_excel(visible_records, parent)
        else:
            logger.warning("Unsupported file format")
            return False
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False